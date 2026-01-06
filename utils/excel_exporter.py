#!/usr/bin/env python3
"""
Excel Exporter for IoTSentinel

Generates professional Excel workbooks with multiple sheets, formatting,
and charts directly from database data.
"""

import io
import sqlite3
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Professional Excel export generator for IoTSentinel.

    Creates multi-sheet workbooks with formatting, formulas, and charts
    directly from database queries.
    """

    def __init__(self, db_path: str):
        """
        Initialize Excel exporter.

        Args:
            db_path: Path to SQLite database
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl not installed. Install with: pip install openpyxl"
            )

        self.db_path = db_path

        # Define color scheme
        self.colors = {
            'header': 'FF2C3E50',
            'subheader': 'FF34495E',
            'success': 'FF27AE60',
            'warning': 'FFF39C12',
            'danger': 'FFE74C3C',
            'info': 'FF3498DB',
            'light_gray': 'FFE8E8E8',
            'white': 'FFFFFFFF'
        }

    def _apply_header_style(self, cell, bg_color='FF2C3E50'):
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color='FFFFFFFF', size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_cell_style(self, cell, bg_color=None):
        """Apply standard cell styling."""
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='FFD3D3D3'),
            right=Side(style='thin', color='FFD3D3D3'),
            top=Side(style='thin', color='FFD3D3D3'),
            bottom=Side(style='thin', color='FFD3D3D3')
        )
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    def _auto_size_columns(self, worksheet):
        """Auto-size columns based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_devices_excel(self) -> bytes:
        """
        Export all devices to Excel format.

        Returns:
            Excel file content as bytes
        """
        try:
            # Query database
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    device_ip,
                    device_name,
                    device_type,
                    mac_address,
                    manufacturer,
                    first_seen,
                    last_seen,
                    is_trusted,
                    is_blocked
                FROM devices
                ORDER BY last_seen DESC
            """)

            devices = cursor.fetchall()
            conn.close()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Device Inventory"

            # Add metadata
            ws['A1'] = 'IoTSentinel Device Inventory Report'
            ws['A1'].font = Font(bold=True, size=16, color=self.colors['header'])
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True, size=10)
            ws['A3'] = f"Total Devices: {len(devices)}"
            ws['A3'].font = Font(bold=True, size=11)

            # Headers
            headers = ['IP Address', 'Device Name', 'Type', 'MAC Address', 'Manufacturer',
                      'Status', 'First Seen', 'Last Seen']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col, value=header)
                self._apply_header_style(cell)

            # Data rows
            row_num = 6
            for device in devices:
                status = 'Blocked' if device['is_blocked'] else 'Trusted' if device['is_trusted'] else 'Active'

                ws.cell(row=row_num, column=1, value=device['device_ip'])
                ws.cell(row=row_num, column=2, value=device['device_name'] or 'Unknown')
                ws.cell(row=row_num, column=3, value=device['device_type'] or 'Unknown')
                ws.cell(row=row_num, column=4, value=device['mac_address'] or 'N/A')
                ws.cell(row=row_num, column=5, value=device['manufacturer'] or 'Unknown')

                # Status cell with color coding
                status_cell = ws.cell(row=row_num, column=6, value=status)
                if status == 'Blocked':
                    status_cell.fill = PatternFill(start_color=self.colors['danger'],
                                                   end_color=self.colors['danger'], fill_type='solid')
                    status_cell.font = Font(color='FFFFFFFF', bold=True)
                elif status == 'Trusted':
                    status_cell.fill = PatternFill(start_color=self.colors['success'],
                                                   end_color=self.colors['success'], fill_type='solid')
                    status_cell.font = Font(color='FFFFFFFF', bold=True)

                ws.cell(row=row_num, column=7, value=self._format_datetime(device['first_seen']))
                ws.cell(row=row_num, column=8, value=self._format_datetime(device['last_seen']))

                # Apply alternating row colors
                bg_color = self.colors['light_gray'] if row_num % 2 == 0 else self.colors['white']
                for col in range(1, 9):
                    if col != 6:  # Skip status column
                        self._apply_cell_style(ws.cell(row=row_num, column=col), bg_color)

                row_num += 1

            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws['A1'] = 'Device Summary Statistics'
            summary_ws['A1'].font = Font(bold=True, size=14)

            # Calculate statistics
            active_count = sum(1 for d in devices if self._is_active(d['last_seen']))
            trusted_count = sum(1 for d in devices if d['is_trusted'])
            blocked_count = sum(1 for d in devices if d['is_blocked'])

            summary_data = [
                ['Metric', 'Count'],
                ['Total Devices', len(devices)],
                ['Active (24h)', active_count],
                ['Trusted', trusted_count],
                ['Blocked', blocked_count],
                ['Inactive', len(devices) - active_count]
            ]

            for row_idx, (metric, count) in enumerate(summary_data, start=3):
                summary_ws.cell(row=row_idx, column=1, value=metric)
                summary_ws.cell(row=row_idx, column=2, value=count)

                if row_idx == 3:
                    self._apply_header_style(summary_ws.cell(row=row_idx, column=1))
                    self._apply_header_style(summary_ws.cell(row=row_idx, column=2))

            # Auto-size columns
            self._auto_size_columns(ws)
            self._auto_size_columns(summary_ws)

            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_data = buffer.getvalue()
            buffer.close()

            logger.info(f"Generated Excel with {len(devices)} devices")
            return excel_data

        except sqlite3.Error as e:
            logger.error(f"Database error exporting devices Excel: {e}")
            return b""
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return b""

    def export_alerts_excel(self, days: int = 7) -> bytes:
        """
        Export alerts to Excel format.

        Args:
            days: Number of days to export (default: 7)

        Returns:
            Excel file content as bytes
        """
        try:
            # Query database
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cutoff_date = datetime.now() - timedelta(days=days)

            cursor.execute("""
                SELECT
                    a.id,
                    a.timestamp,
                    a.device_ip,
                    d.device_name,
                    a.severity,
                    a.anomaly_score,
                    a.explanation,
                    a.acknowledged
                FROM alerts a
                LEFT JOIN devices d ON a.device_ip = d.device_ip
                WHERE a.timestamp > ?
                ORDER BY a.timestamp DESC
            """, (cutoff_date.isoformat(),))

            alerts = cursor.fetchall()
            conn.close()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Alerts"

            # Add metadata
            ws['A1'] = 'IoTSentinel Security Alerts Report'
            ws['A1'].font = Font(bold=True, size=16, color=self.colors['header'])
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True, size=10)
            ws['A3'] = f"Period: Last {days} days | Total Alerts: {len(alerts)}"
            ws['A3'].font = Font(bold=True, size=11)

            # Headers
            headers = ['Alert ID', 'Timestamp', 'Device IP', 'Device Name',
                      'Severity', 'Anomaly Score', 'Explanation', 'Acknowledged']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col, value=header)
                self._apply_header_style(cell)

            # Data rows
            row_num = 6
            for alert in alerts:
                ws.cell(row=row_num, column=1, value=alert['id'])
                ws.cell(row=row_num, column=2, value=self._format_datetime(alert['timestamp']))
                ws.cell(row=row_num, column=3, value=alert['device_ip'])
                ws.cell(row=row_num, column=4, value=alert['device_name'] or 'Unknown')

                # Severity with color coding
                severity_cell = ws.cell(row=row_num, column=5, value=alert['severity'].upper())
                sev_lower = alert['severity'].lower()
                if sev_lower == 'critical':
                    color = self.colors['danger']
                elif sev_lower == 'high':
                    color = self.colors['warning']
                elif sev_lower == 'medium':
                    color = self.colors['info']
                else:
                    color = self.colors['success']
                severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                severity_cell.font = Font(color='FFFFFFFF', bold=True)

                ws.cell(row=row_num, column=6, value=alert['anomaly_score'] or 0)
                ws.cell(row=row_num, column=7, value=alert['explanation'] or '')
                ws.cell(row=row_num, column=8, value='Yes' if alert['acknowledged'] else 'No')

                # Apply alternating row colors
                bg_color = self.colors['light_gray'] if row_num % 2 == 0 else self.colors['white']
                for col in [1, 2, 3, 4, 6, 7, 8]:  # Skip severity column
                    self._apply_cell_style(ws.cell(row=row_num, column=col), bg_color)

                row_num += 1

            # Add summary sheet
            summary_ws = wb.create_sheet("Severity Summary")
            summary_ws['A1'] = 'Alert Severity Breakdown'
            summary_ws['A1'].font = Font(bold=True, size=14)

            # Calculate severity statistics
            severity_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
            for alert in alerts:
                sev = alert['severity'].lower()
                if sev in severity_counts:
                    severity_counts[sev] += 1

            summary_data = [
                ['Severity', 'Count', 'Percentage'],
                ['Critical', severity_counts['critical'],
                 f"{severity_counts['critical']/len(alerts)*100:.1f}%" if alerts else "0%"],
                ['High', severity_counts['high'],
                 f"{severity_counts['high']/len(alerts)*100:.1f}%" if alerts else "0%"],
                ['Medium', severity_counts['medium'],
                 f"{severity_counts['medium']/len(alerts)*100:.1f}%" if alerts else "0%"],
                ['Low', severity_counts['low'],
                 f"{severity_counts['low']/len(alerts)*100:.1f}%" if alerts else "0%"]
            ]

            for row_idx, row_data in enumerate(summary_data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 3:
                        self._apply_header_style(cell)

            # Auto-size columns
            self._auto_size_columns(ws)
            self._auto_size_columns(summary_ws)

            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_data = buffer.getvalue()
            buffer.close()

            logger.info(f"Generated Excel with {len(alerts)} alerts")
            return excel_data

        except sqlite3.Error as e:
            logger.error(f"Database error exporting alerts Excel: {e}")
            return b""
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return b""

    def export_connections_excel(self, device_ip: Optional[str] = None, hours: int = 24) -> bytes:
        """
        Export connection logs to Excel format.

        Args:
            device_ip: Filter by specific device IP (optional)
            hours: Number of hours to export (default: 24)

        Returns:
            Excel file content as bytes
        """
        try:
            # Query database
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cutoff_time = datetime.now() - timedelta(hours=hours)

            if device_ip:
                cursor.execute("""
                    SELECT
                        timestamp,
                        device_ip,
                        dest_ip,
                        dest_port,
                        protocol,
                        service,
                        bytes_sent,
                        bytes_received,
                        packets_sent,
                        packets_received,
                        conn_state
                    FROM connections
                    WHERE device_ip = ? AND timestamp > ?
                    ORDER BY timestamp DESC
                    LIMIT 5000
                """, (device_ip, cutoff_time.isoformat()))
            else:
                cursor.execute("""
                    SELECT
                        timestamp,
                        device_ip,
                        dest_ip,
                        dest_port,
                        protocol,
                        service,
                        bytes_sent,
                        bytes_received,
                        packets_sent,
                        packets_received,
                        conn_state
                    FROM connections
                    WHERE timestamp > ?
                    ORDER BY timestamp DESC
                    LIMIT 5000
                """, (cutoff_time.isoformat(),))

            connections = cursor.fetchall()
            conn.close()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Connections"

            # Add metadata
            title_text = 'Network Connections Report'
            if device_ip:
                title_text += f' - {device_ip}'
            ws['A1'] = title_text
            ws['A1'].font = Font(bold=True, size=16, color=self.colors['header'])
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True, size=10)
            ws['A3'] = f"Time Range: Last {hours} hours | Total Connections: {len(connections)}"
            ws['A3'].font = Font(bold=True, size=11)

            # Headers
            headers = ['Timestamp', 'Source IP', 'Destination IP', 'Port', 'Protocol',
                      'Service', 'Bytes Sent', 'Bytes Received', 'Packets Sent',
                      'Packets Received', 'State']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col, value=header)
                self._apply_header_style(cell, self.colors['success'])

            # Data rows
            row_num = 6
            for conn in connections:
                ws.cell(row=row_num, column=1, value=self._format_datetime(conn['timestamp']))
                ws.cell(row=row_num, column=2, value=conn['device_ip'] or 'N/A')
                ws.cell(row=row_num, column=3, value=conn['dest_ip'] or 'N/A')
                ws.cell(row=row_num, column=4, value=conn['dest_port'] or 'N/A')
                ws.cell(row=row_num, column=5, value=conn['protocol'] or 'N/A')
                ws.cell(row=row_num, column=6, value=conn['service'] or 'Unknown')
                ws.cell(row=row_num, column=7, value=conn['bytes_sent'] or 0)
                ws.cell(row=row_num, column=8, value=conn['bytes_received'] or 0)
                ws.cell(row=row_num, column=9, value=conn['packets_sent'] or 0)
                ws.cell(row=row_num, column=10, value=conn['packets_received'] or 0)
                ws.cell(row=row_num, column=11, value=conn['conn_state'] or 'Unknown')

                # Apply alternating row colors
                bg_color = self.colors['light_gray'] if row_num % 2 == 0 else self.colors['white']
                for col in range(1, 12):
                    self._apply_cell_style(ws.cell(row=row_num, column=col), bg_color)

                row_num += 1

            # Auto-size columns
            self._auto_size_columns(ws)

            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_data = buffer.getvalue()
            buffer.close()

            logger.info(f"Generated Excel with {len(connections)} connections")
            return excel_data

        except sqlite3.Error as e:
            logger.error(f"Database error exporting connections Excel: {e}")
            return b""
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return b""

    # Helper methods

    def _is_active(self, last_seen: str, hours: int = 24) -> bool:
        """Check if device was active within specified hours."""
        if not last_seen:
            return False
        try:
            last_seen_dt = datetime.fromisoformat(last_seen)
            cutoff = datetime.now() - timedelta(hours=hours)
            return last_seen_dt > cutoff
        except:
            return False

    def _format_datetime(self, dt_str: str) -> str:
        """Format datetime string for display."""
        if not dt_str:
            return 'N/A'
        try:
            dt = datetime.fromisoformat(dt_str)
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return dt_str
