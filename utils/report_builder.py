#!/usr/bin/env python3
"""
Custom Report Builder for IoTSentinel

Builds reports based on templates by executing sections and compiling results.
Integrates TrendAnalyzer, ChartFactory, and ReportGenerator.
"""

import logging
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
import json
from io import BytesIO

from .report_templates import ReportTemplateManager, ReportTemplate, ReportSection
from .trend_analyzer import TrendAnalyzer
from .chart_factory import ChartFactory
from .report_generator import ReportGenerator
from .report_cache import ReportCache

logger = logging.getLogger(__name__)


class ReportBuilder:
    """
    Builds custom reports based on templates.

    Executes template sections, fetches data, generates visualizations,
    and compiles final report in requested format.
    """

    def __init__(
        self,
        db_path: str,
        enable_cache: bool = True,
        cache_ttl_minutes: int = 15
    ):
        """
        Initialize report builder.

        Args:
            db_path: Path to SQLite database
            enable_cache: Enable report caching
            cache_ttl_minutes: Cache TTL in minutes
        """
        self.db_path = db_path
        self.template_manager = ReportTemplateManager()
        self.trend_analyzer = TrendAnalyzer(db_path)
        self.report_generator = ReportGenerator(db_path)

        # Initialize cache
        self.enable_cache = enable_cache
        if enable_cache:
            self.cache = ReportCache(
                cache_dir='data/cache/reports',
                default_ttl_minutes=cache_ttl_minutes,
                max_cache_size_mb=100
            )
            logger.info("Report caching enabled")
        else:
            self.cache = None

    def build_report(
        self,
        template_name: str,
        format: str = 'json',
        parameters: Optional[Dict[str, Any]] = None,
        use_cache: bool = True
    ) -> Optional[Dict[str, Any]]:
        """
        Build a report from a template with caching support.

        Args:
            template_name: Name of the template to use
            format: Output format ('json', 'pdf', 'excel', 'html')
            parameters: Optional parameters to override template defaults
            use_cache: Whether to use caching (default True)

        Returns:
            Dictionary with report content and metadata
        """
        try:
            params = parameters or {}

            # Check cache first
            if self.enable_cache and self.cache and use_cache:
                cached_report = self.cache.get(template_name, format, params)
                if cached_report:
                    logger.info(f"Using cached report: {template_name} ({format})")
                    return cached_report

            # Get template
            template = self.template_manager.get_template(template_name)
            if not template:
                logger.error(f"Template not found: {template_name}")
                return None

            # Execute all sections
            sections_data = self._execute_sections(template, params)

            # Compile report in requested format
            report_data = None
            if format == 'json':
                report_data = self._compile_json_report(template, sections_data)
            elif format == 'pdf':
                report_data = self._compile_pdf_report(template, sections_data)
            elif format == 'excel':
                report_data = self._compile_excel_report(template, sections_data)
            elif format == 'html':
                report_data = self._compile_html_report(template, sections_data)
            else:
                logger.error(f"Unsupported format: {format}")
                return None

            # Cache the generated report
            if report_data and self.enable_cache and self.cache and use_cache:
                # Determine cache TTL based on time range in parameters
                days = params.get('days', 7)
                # Shorter TTL for recent data (changes frequently)
                # Longer TTL for historical data (more stable)
                if days <= 1:
                    ttl_minutes = 5  # 5 minutes for daily reports
                elif days <= 7:
                    ttl_minutes = 15  # 15 minutes for weekly reports
                elif days <= 30:
                    ttl_minutes = 60  # 1 hour for monthly reports
                else:
                    ttl_minutes = 180  # 3 hours for longer periods

                self.cache.put(
                    template_name,
                    format,
                    params,
                    report_data,
                    ttl_minutes=ttl_minutes
                )
                logger.info(f"Cached report: {template_name} ({format}, TTL: {ttl_minutes}min)")

            return report_data

        except Exception as e:
            logger.error(f"Error building report: {e}")
            return None

    def _execute_sections(
        self,
        template: ReportTemplate,
        parameters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Execute all sections in the template.

        Args:
            template: Report template
            parameters: Parameters for execution

        Returns:
            List of executed section data
        """
        sections_data = []

        for section in sorted(template.sections, key=lambda s: s.order):
            try:
                section_data = self._execute_section(section, parameters)
                sections_data.append(section_data)
            except Exception as e:
                logger.error(f"Error executing section '{section.title}': {e}")
                sections_data.append({
                    'title': section.title,
                    'error': str(e),
                    'section_type': section.section_type
                })

        return sections_data

    def _execute_section(
        self,
        section: ReportSection,
        parameters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Execute a single section.

        Args:
            section: Section to execute
            parameters: Execution parameters

        Returns:
            Section data with results
        """
        section_data = {
            'title': section.title,
            'section_type': section.section_type,
            'data_source': section.data_source
        }

        if section.section_type == 'metrics':
            section_data['content'] = self._get_metrics_data(section, parameters)
        elif section.section_type == 'chart':
            section_data['content'] = self._get_chart_data(section, parameters)
        elif section.section_type == 'table':
            section_data['content'] = self._get_table_data(section, parameters)
        elif section.section_type == 'list':
            section_data['content'] = self._get_list_data(section, parameters)
        elif section.section_type == 'text':
            section_data['content'] = self._get_text_data(section, parameters)
        else:
            section_data['content'] = {'error': f'Unknown section type: {section.section_type}'}

        return section_data

    def _get_metrics_data(
        self,
        section: ReportSection,
        parameters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Get metrics data for a section."""
        config = section.config
        metrics = {}

        if section.data_source == 'alerts':
            days = parameters.get('days', config.get('time_period', 7))
            summary = self.trend_analyzer.get_executive_summary(days=days)
            security_posture = summary.get('security_posture', {})

            for metric in config.get('metrics', []):
                metrics[metric] = security_posture.get(metric, 'N/A')

        elif section.data_source == 'connections':
            hours = parameters.get('hours', config.get('time_period', 24))
            traffic = self.trend_analyzer.analyze_network_traffic(hours=hours)

            for metric in config.get('metrics', []):
                metrics[metric] = traffic.get(metric, 'N/A')

        elif section.data_source == 'devices':
            stats = self.report_generator.get_summary_statistics(days=7)
            for metric in config.get('metrics', []):
                metrics[metric] = stats.get(metric, 'N/A')

        return metrics

    def _get_chart_data(
        self,
        section: ReportSection,
        parameters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Get chart data for a section."""
        config = section.config
        chart_type = config.get('chart_type', 'bar')

        if section.data_source == 'alerts':
            days = parameters.get('days', config.get('period_days', 7))

            if chart_type == 'trend' or chart_type == 'area':
                trend_data = self.trend_analyzer.analyze_alert_trends(
                    days=days,
                    granularity='daily'
                )
                time_series = trend_data.get('time_series', [])

                if time_series:
                    x_vals = [t[0] for t in time_series]
                    y_vals = [t[1] for t in time_series]

                    if chart_type == 'trend':
                        return ChartFactory.create_trend_chart(
                            x_vals, y_vals,
                            title=section.title,
                            x_title='Date',
                            y_title='Alert Count'
                        )
                    else:  # area
                        return ChartFactory.create_area_chart(
                            x_vals, y_vals,
                            title=section.title,
                            x_title='Date',
                            y_title='Alerts'
                        )

            elif chart_type == 'pie':
                # Alert severity distribution
                days = parameters.get('days', config.get('period_days', 7))
                stats = self.report_generator.get_summary_statistics(days=days)
                by_severity = stats.get('alerts_by_severity', {})

                if by_severity:
                    labels = list(by_severity.keys())
                    values = list(by_severity.values())
                    return ChartFactory.create_pie_chart(
                        labels, values,
                        title=section.title
                    )

        elif section.data_source == 'devices':
            if chart_type == 'bar':
                days = parameters.get('days', 7)
                activity = self.trend_analyzer.analyze_device_activity(days=days)
                devices = activity.get('most_active_devices', [])[:10]

                if devices:
                    ips = [d[0] for d in devices]
                    counts = [d[1] for d in devices]
                    return ChartFactory.create_bar_chart(
                        ips, counts,
                        title=section.title,
                        x_title='Device IP',
                        y_title='Connection Count',
                        tick_angle=-30
                    )

        return ChartFactory.create_empty_chart('No data available')

    def _get_table_data(
        self,
        section: ReportSection,
        parameters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Get table data for a section."""
        import sqlite3

        config = section.config
        conn = self.db_manager.conn

        cursor = conn.cursor()

        try:
            if section.data_source == 'alerts':
                days = parameters.get('days', config.get('period_days', 7))
                cutoff = datetime.now() - timedelta(days=days)

                # Build query based on filters
                query = "SELECT * FROM alerts WHERE timestamp >= ?"
                params = [cutoff.isoformat()]

                filter_config = config.get('filter', {})
                if 'severity' in filter_config:
                    severity = filter_config['severity']
                    if isinstance(severity, list):
                        placeholders = ','.join('?' * len(severity))
                        query += f" AND severity IN ({placeholders})"
                        params.extend(severity)
                    else:
                        query += " AND severity = ?"
                        params.append(severity)

                query += " ORDER BY timestamp DESC"

                if 'limit' in config:
                    query += f" LIMIT {config['limit']}"

                cursor.execute(query, params)

            elif section.data_source == 'devices':
                query = "SELECT * FROM devices"
                params = []

                filter_config = config.get('filter', {})
                if filter_config:
                    conditions = []
                    if 'is_blocked' in filter_config:
                        conditions.append("is_blocked = ?")
                        params.append(1 if filter_config['is_blocked'] else 0)

                    if conditions:
                        query += " WHERE " + " AND ".join(conditions)

                if 'sort_by' in config:
                    query += f" ORDER BY {config['sort_by']}"
                    if config.get('sort_order') == 'desc':
                        query += " DESC"

                if 'limit' in config:
                    query += f" LIMIT {config['limit']}"

                cursor.execute(query, params)

            elif section.data_source == 'connections':
                hours = parameters.get('hours', config.get('period_hours', 24))
                cutoff = datetime.now() - timedelta(hours=hours)

                query = "SELECT * FROM connections WHERE timestamp >= ?"
                params = [cutoff.isoformat()]

                query += " ORDER BY timestamp DESC"

                if 'limit' in config:
                    query += f" LIMIT {config['limit']}"

                cursor.execute(query, params)

            rows = cursor.fetchall()

            # Convert to list of dicts
            table_data = {
                'columns': config.get('columns', [col[0] for col in cursor.description]),
                'rows': [dict(row) for row in rows],
                'row_count': len(rows)
            }

            return table_data

        except Exception as e:
            logger.error(f"Error getting table data: {e}")
            return {'error': str(e)}

    def _get_list_data(
        self,
        section: ReportSection,
        parameters: Dict[str, Any]
    ) -> List[str]:
        """Get list data for a section."""
        config = section.config

        if section.data_source == 'custom':
            source = config.get('source', '')

            if source == 'trend_analyzer.get_executive_summary':
                days = parameters.get('days', 7)
                summary = self.trend_analyzer.get_executive_summary(days=days)
                field = config.get('field', 'top_concerns')
                return summary.get(field, [])

        return []

    def _get_text_data(
        self,
        section: ReportSection,
        parameters: Dict[str, Any]
    ) -> str:
        """Get text data for a section."""
        config = section.config
        content_type = config.get('content_type', 'static')

        if content_type == 'recommendations':
            # Generate security recommendations based on data
            days = parameters.get('days', 7)
            summary = self.trend_analyzer.get_executive_summary(days=days)
            concerns = summary.get('top_concerns', [])

            recommendations = ["Based on the analysis, we recommend:"]
            if any('increasing' in str(c).lower() for c in concerns):
                recommendations.append("- Review and strengthen alert rules to reduce false positives")
            if any('suspicious' in str(c).lower() for c in concerns):
                recommendations.append("- Investigate suspicious network patterns immediately")
            if any('new devices' in str(c).lower() for c in concerns):
                recommendations.append("- Verify authorization for all new devices detected")

            return "\n".join(recommendations)

        return config.get('text', '')

    def _compile_json_report(
        self,
        template: ReportTemplate,
        sections_data: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Compile report as JSON."""
        report = {
            'template': template.name,
            'generated_at': datetime.now().isoformat(),
            'sections': sections_data
        }

        return {
            'content': json.dumps(report, indent=2, default=str),
            'filename': f'{template.name.lower().replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',
            'mimetype': 'application/json'
        }

    def _compile_pdf_report(
        self,
        template: ReportTemplate,
        sections_data: List[Dict[str, Any]]
    ) -> Optional[Dict[str, Any]]:
        """Compile report as PDF."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#0d6efd'),
                spaceAfter=30
            )
            story.append(Paragraph(template.name, title_style))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))

            # Process each section
            for section in sections_data:
                # Section title
                story.append(Paragraph(f"<b>{section['title']}</b>", styles['Heading2']))
                story.append(Spacer(1, 0.1*inch))

                # Section content
                if section.get('error'):
                    story.append(Paragraph(f"Error: {section['error']}", styles['Normal']))
                elif section['section_type'] == 'metrics':
                    # Render metrics as table
                    metrics = section.get('content', {})
                    if metrics:
                        data = [['Metric', 'Value']]
                        for key, value in metrics.items():
                            data.append([key.replace('_', ' ').title(), str(value)])

                        table = Table(data, colWidths=[3*inch, 2*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)

                elif section['section_type'] == 'table':
                    table_data = section.get('content', {})
                    if not table_data.get('error') and table_data.get('rows'):
                        # Limit rows for PDF
                        rows = table_data['rows'][:50]
                        columns = table_data.get('columns', [])

                        data = [columns]
                        for row in rows:
                            data.append([str(row.get(col, ''))[:50] for col in columns])

                        table = Table(data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                            ('FONTSIZE', (0, 0), (-1, -1), 8)
                        ]))
                        story.append(table)
                        if len(table_data['rows']) > 50:
                            story.append(Paragraph(f"(Showing 50 of {len(table_data['rows'])} rows)", styles['Italic']))

                elif section['section_type'] == 'list':
                    items = section.get('content', [])
                    for item in items:
                        story.append(Paragraph(f"• {item}", styles['Normal']))
                        story.append(Spacer(1, 0.05*inch))

                elif section['section_type'] == 'text':
                    text = section.get('content', '')
                    story.append(Paragraph(text, styles['Normal']))

                story.append(Spacer(1, 0.3*inch))

            # Build PDF
            doc.build(story)
            pdf_content = buffer.getvalue()
            buffer.close()

            return {
                'content': pdf_content,
                'filename': f'{template.name.lower().replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
                'mimetype': 'application/pdf'
            }

        except Exception as e:
            logger.error(f"Error compiling PDF report: {e}")
            return None

    def _compile_excel_report(
        self,
        template: ReportTemplate,
        sections_data: List[Dict[str, Any]]
    ) -> Optional[Dict[str, Any]]:
        """Compile report as Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create a sheet for each section
            for section in sections_data:
                sheet_name = section['title'][:31]  # Excel limit
                ws = wb.create_sheet(sheet_name)

                # Write section title
                ws['A1'] = section['title']
                ws['A1'].font = Font(size=14, bold=True)

                row = 3

                if section.get('error'):
                    ws[f'A{row}'] = f"Error: {section['error']}"
                elif section['section_type'] == 'metrics':
                    metrics = section.get('content', {})
                    ws[f'A{row}'] = 'Metric'
                    ws[f'B{row}'] = 'Value'
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'].font = Font(bold=True)
                    row += 1

                    for key, value in metrics.items():
                        ws[f'A{row}'] = key.replace('_', ' ').title()
                        ws[f'B{row}'] = value
                        row += 1

                elif section['section_type'] == 'table':
                    table_data = section.get('content', {})
                    if not table_data.get('error') and table_data.get('rows'):
                        columns = table_data.get('columns', [])

                        # Write headers
                        for col_idx, col_name in enumerate(columns, 1):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.value = col_name
                            cell.font = Font(bold=True)

                        row += 1

                        # Write data
                        for data_row in table_data['rows']:
                            for col_idx, col_name in enumerate(columns, 1):
                                ws.cell(row=row, column=col_idx, value=str(data_row.get(col_name, '')))
                            row += 1

                elif section['section_type'] == 'list':
                    items = section.get('content', [])
                    for item in items:
                        ws[f'A{row}'] = item
                        row += 1

            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()

            return {
                'content': excel_content,
                'filename': f'{template.name.lower().replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }

        except Exception as e:
            logger.error(f"Error compiling Excel report: {e}")
            return None

    def _compile_html_report(
        self,
        template: ReportTemplate,
        sections_data: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Compile report as HTML."""
        html_parts = [
            '<!DOCTYPE html>',
            '<html>',
            '<head>',
            f'<title>{template.name}</title>',
            '<style>',
            'body { font-family: Arial, sans-serif; margin: 20px; }',
            'h1 { color: #0d6efd; }',
            'h2 { color: #333; border-bottom: 2px solid #0d6efd; padding-bottom: 5px; }',
            'table { border-collapse: collapse; width: 100%; margin: 10px 0; }',
            'th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }',
            'th { background-color: #0d6efd; color: white; }',
            'tr:nth-child(even) { background-color: #f2f2f2; }',
            '.metrics { background-color: #f8f9fa; padding: 15px; border-radius: 5px; }',
            '</style>',
            '</head>',
            '<body>',
            f'<h1>{template.name}</h1>',
            f'<p>Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>'
        ]

        for section in sections_data:
            html_parts.append(f'<h2>{section["title"]}</h2>')

            if section.get('error'):
                html_parts.append(f'<p style="color: red;">Error: {section["error"]}</p>')
            elif section['section_type'] == 'metrics':
                metrics = section.get('content', {})
                html_parts.append('<div class="metrics">')
                for key, value in metrics.items():
                    html_parts.append(f'<p><strong>{key.replace("_", " ").title()}:</strong> {value}</p>')
                html_parts.append('</div>')

            elif section['section_type'] == 'table':
                table_data = section.get('content', {})
                if not table_data.get('error') and table_data.get('rows'):
                    columns = table_data.get('columns', [])
                    html_parts.append('<table>')
                    html_parts.append('<tr>')
                    for col in columns:
                        html_parts.append(f'<th>{col}</th>')
                    html_parts.append('</tr>')

                    for row in table_data['rows'][:100]:  # Limit rows
                        html_parts.append('<tr>')
                        for col in columns:
                            html_parts.append(f'<td>{row.get(col, "")}</td>')
                        html_parts.append('</tr>')

                    html_parts.append('</table>')

            elif section['section_type'] == 'list':
                items = section.get('content', [])
                html_parts.append('<ul>')
                for item in items:
                    html_parts.append(f'<li>{item}</li>')
                html_parts.append('</ul>')

        html_parts.extend(['</body>', '</html>'])

        return {
            'content': '\n'.join(html_parts),
            'filename': f'{template.name.lower().replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html',
            'mimetype': 'text/html'
        }

    def list_available_templates(self) -> List[Dict[str, str]]:
        """List all available templates."""
        return self.template_manager.list_templates()

    def get_template_preview(self, template_name: str) -> Optional[Dict[str, Any]]:
        """Get a preview of template structure."""
        return self.template_manager.get_template_config(template_name)

    # Cache Management Methods

    def get_cache_stats(self) -> Optional[Dict[str, Any]]:
        """
        Get cache statistics.

        Returns:
            Dictionary with cache statistics or None if caching disabled
        """
        if self.enable_cache and self.cache:
            return self.cache.get_stats()
        return None

    def clear_cache(self):
        """Clear all cached reports."""
        if self.enable_cache and self.cache:
            self.cache.invalidate()
            logger.info("Report cache cleared")
        else:
            logger.warning("Caching is not enabled")

    def clear_expired_cache(self):
        """Clear only expired cache entries."""
        if self.enable_cache and self.cache:
            self.cache.clear_expired()
            logger.info("Expired cache entries cleared")
        else:
            logger.warning("Caching is not enabled")

    def invalidate_template_cache(self, template_name: str):
        """
        Invalidate cache for specific template.

        Args:
            template_name: Template to invalidate
        """
        if self.enable_cache and self.cache:
            self.cache.invalidate(template_name=template_name)
            logger.info(f"Cache invalidated for template: {template_name}")
        else:
            logger.warning("Caching is not enabled")
