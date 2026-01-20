#!/usr/bin/env python3
"""
Export Helpers for Dashboard Integration

Helper functions to integrate UniversalExporter with Dash callbacks.
Provides simplified interfaces for common export operations.
"""

import logging
from typing import Dict, Any, Optional
from .universal_exporter import UniversalExporter, ExportFormat

logger = logging.getLogger(__name__)


class DashExportHelper:
    """
    Helper class for integrating exports into Dash callbacks.

    Provides simple methods that return data in the format expected
    by dcc.Download components.
    """

    def __init__(self, db_path: str = None, db_manager=None):
        """
        Initialize export helper.

        Args:
            db_path: Path to SQLite database (legacy)
            db_manager: DatabaseManager instance (preferred)
        """
        self.exporter = UniversalExporter(db_path=db_path, db_manager=db_manager)

    def export_devices(
        self,
        format: str = 'csv'
    ) -> Optional[Dict[str, Any]]:
        """
        Export devices in specified format.

        Args:
            format: One of 'csv', 'json', 'pdf', 'excel'

        Returns:
            Download dict for dcc.Download or None on error
        """
        try:
            result = self.exporter.export_devices(format=format)

            if not result or not result.get('content'):
                return None

            # Handle binary content (PDF, Excel)
            if isinstance(result['content'], bytes):
                import base64
                content_b64 = base64.b64encode(result['content']).decode()
                return {
                    'content': content_b64,
                    'filename': result['filename'],
                    'type': result['mimetype'],
                    'base64': True
                }
            else:
                # Text content (CSV, JSON)
                return {
                    'content': result['content'],
                    'filename': result['filename'],
                    'type': result['mimetype']
                }

        except Exception as e:
            logger.error(f"Error in export_devices: {e}")
            return None

    def export_alerts(
        self,
        format: str = 'csv',
        days: int = 7
    ) -> Optional[Dict[str, Any]]:
        """
        Export alerts in specified format.

        Args:
            format: One of 'csv', 'json', 'pdf', 'excel'
            days: Number of days to include

        Returns:
            Download dict for dcc.Download or None on error
        """
        try:
            result = self.exporter.export_alerts(format=format, days=days)

            if not result or not result.get('content'):
                return None

            # Handle binary content (PDF, Excel)
            if isinstance(result['content'], bytes):
                import base64
                content_b64 = base64.b64encode(result['content']).decode()
                return {
                    'content': content_b64,
                    'filename': result['filename'],
                    'type': result['mimetype'],
                    'base64': True
                }
            else:
                # Text content (CSV, JSON)
                return {
                    'content': result['content'],
                    'filename': result['filename'],
                    'type': result['mimetype']
                }

        except Exception as e:
            logger.error(f"Error in export_alerts: {e}")
            return None

    def export_connections(
        self,
        format: str = 'csv',
        device_ip: Optional[str] = None,
        hours: int = 24
    ) -> Optional[Dict[str, Any]]:
        """
        Export connections in specified format.

        Args:
            format: One of 'csv', 'json', 'pdf', 'excel'
            device_ip: Filter by device IP (optional)
            hours: Number of hours to include

        Returns:
            Download dict for dcc.Download or None on error
        """
        try:
            result = self.exporter.export_connections(
                format=format,
                device_ip=device_ip,
                hours=hours
            )

            if not result or not result.get('content'):
                return None

            # Handle binary content (PDF, Excel)
            if isinstance(result['content'], bytes):
                import base64
                content_b64 = base64.b64encode(result['content']).decode()
                return {
                    'content': content_b64,
                    'filename': result['filename'],
                    'type': result['mimetype'],
                    'base64': True
                }
            else:
                # Text content (CSV, JSON)
                return {
                    'content': result['content'],
                    'filename': result['filename'],
                    'type': result['mimetype']
                }

        except Exception as e:
            logger.error(f"Error in export_connections: {e}")
            return None

    def export_sustainability(
        self,
        format: str = 'csv',
        carbon_data: dict = None,
        energy_data: dict = None
    ) -> Optional[Dict[str, Any]]:
        """
        Export sustainability metrics in specified format.

        Args:
            format: One of 'csv', 'json', 'pdf', 'excel'
            carbon_data: Carbon footprint metrics dict
            energy_data: Energy consumption metrics dict

        Returns:
            Download dict for dcc.Download or None on error
        """
        try:
            from datetime import datetime

            if not carbon_data or not energy_data:
                return None

            # Use the universal exporter's methods with sustainability data
            result = self.exporter.export_sustainability_report(
                format=format,
                carbon_data=carbon_data,
                energy_data=energy_data
            )

            if not result or not result.get('content'):
                return None

            # Handle binary content (PDF, Excel)
            if isinstance(result['content'], bytes):
                import base64
                content_b64 = base64.b64encode(result['content']).decode()
                return {
                    'content': content_b64,
                    'filename': result['filename'],
                    'type': result['mimetype'],
                    'base64': True
                }
            else:
                # Text content (CSV, JSON)
                return {
                    'content': result['content'],
                    'filename': result['filename'],
                    'type': result['mimetype']
                }

        except Exception as e:
            logger.error(f"Error in export_sustainability: {e}")
            return None

    def export_integrations(
        self,
        format: str = 'json'
    ) -> Optional[Dict[str, Any]]:
        """
        Export API Integration Hub configuration in specified format.

        Args:
            format: One of 'json', 'csv', 'pdf', 'excel'

        Returns:
            Download dict for dcc.Download or None on error
        """
        try:
            from datetime import datetime
            from alerts.integration_system import IntegrationManager
            import json
            import csv
            from io import StringIO, BytesIO

            mgr = IntegrationManager(self.exporter.db_manager)
            integrations = mgr.get_all_integrations()

            # Build export data (exclude encrypted credentials)
            export_data = {
                'export_timestamp': datetime.now().isoformat(),
                'total_integrations': len(integrations),
                'enabled_count': len([i for i in integrations if i.get('is_enabled')]),
                'integrations': []
            }

            for integration in integrations:
                export_data['integrations'].append({
                    'id': integration['id'],
                    'name': integration['name'],
                    'category': integration['category'],
                    'is_enabled': integration.get('is_enabled', False),
                    'health_status': integration.get('health_status', 'untested'),
                    'total_requests': integration.get('total_requests', 0),
                    'successful_requests': integration.get('successful_requests', 0),
                    'failed_requests': integration.get('failed_requests', 0),
                    'last_used': integration.get('last_used'),
                    'last_health_check': integration.get('last_health_check')
                    # Note: Credentials NOT exported for security
                })

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            if format == 'json':
                content = json.dumps(export_data, indent=2)
                filename = f"api_hub_config_{timestamp}.json"
                mimetype = 'application/json'

            elif format == 'csv':
                # CSV format - flatten integrations list
                output = StringIO()
                if export_data['integrations']:
                    fieldnames = export_data['integrations'][0].keys()
                    writer = csv.DictWriter(output, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(export_data['integrations'])
                    content = output.getvalue()
                else:
                    content = "No integrations configured\n"

                filename = f"api_hub_config_{timestamp}.csv"
                mimetype = 'text/csv'

            elif format == 'pdf':
                # PDF format using ReportLab
                from reportlab.lib.pagesizes import letter
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                styles = getSampleStyleSheet()
                elements = []

                # Title
                title = Paragraph("<b>API Integration Hub Configuration</b>", styles['Title'])
                elements.append(title)
                elements.append(Spacer(1, 12))

                # Metadata
                timestamp_para = Paragraph(
                    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
                    f"Total Integrations: {export_data['total_integrations']}<br/>"
                    f"Enabled: {export_data['enabled_count']}<br/>"
                    f"<i>Note: Credentials excluded for security</i>",
                    styles['Normal']
                )
                elements.append(timestamp_para)
                elements.append(Spacer(1, 20))

                # Integration Table
                table_data = [['Name', 'Category', 'Enabled', 'Health', 'Requests']]
                for integration in export_data['integrations']:
                    table_data.append([
                        integration['name'],
                        integration['category'].replace('_', ' ').title(),
                        'Yes' if integration['is_enabled'] else 'No',
                        integration['health_status'].title(),
                        str(integration['total_requests'])
                    ])

                table = Table(table_data, colWidths=[120, 100, 60, 80, 70])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8)
                ]))
                elements.append(table)

                doc.build(elements)
                content = buffer.getvalue()
                filename = f"api_hub_config_{timestamp}.pdf"
                mimetype = 'application/pdf'

            elif format == 'excel':
                # Excel format using openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "API Integrations"

                # Header row
                headers = ['Name', 'Category', 'Enabled', 'Health Status', 'Total Requests',
                          'Successful', 'Failed', 'Last Used', 'Last Health Check']
                ws.append(headers)

                # Style header
                header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='left')

                # Data rows
                for integration in export_data['integrations']:
                    ws.append([
                        integration['name'],
                        integration['category'].replace('_', ' ').title(),
                        'Yes' if integration['is_enabled'] else 'No',
                        integration['health_status'].title(),
                        integration['total_requests'],
                        integration['successful_requests'],
                        integration['failed_requests'],
                        integration['last_used'] or 'Never',
                        integration['last_health_check'] or 'Never'
                    ])

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                buffer = BytesIO()
                wb.save(buffer)
                content = buffer.getvalue()
                filename = f"api_hub_config_{timestamp}.xlsx"
                mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            else:
                logger.warning(f"Unsupported format for integrations export: {format}")
                return None

            # Return in format expected by dcc.Download
            if format in ['pdf', 'excel']:
                # Binary content - need base64 encoding
                import base64
                result = {
                    'content': base64.b64encode(content).decode(),
                    'filename': filename,
                    'type': mimetype,
                    'base64': True
                }
            else:
                # Text content (JSON, CSV)
                result = {
                    'content': content,
                    'filename': filename,
                    'type': mimetype
                }

            logger.info(f"Successfully prepared integration export: {filename}")
            return result

        except Exception as e:
            logger.error(f"Error in export_integrations: {e}", exc_info=True)
            return None
            logger.error(f"Error in export_integrations: {e}")
            return None


# Example Dash callback implementations
"""
EXAMPLE USAGE IN app.py:

1. Import the helper at the top of app.py:
   from utils.export_helpers import DashExportHelper

2. Initialize the helper (after DB_PATH is defined):
   export_helper = DashExportHelper(DB_PATH)

3. Create a universal export callback:

@app.callback(
    [Output('download-devices', 'data'),
     Output('devices-export-toast', 'children')],
    [Input('export-devices-btn', 'n_clicks')],
    [State('export-format-dropdown', 'value')]
)
def export_devices_universal(n_clicks, export_format):
    '''Export devices in selected format (CSV/JSON/PDF/Excel).'''
    if not n_clicks:
        raise dash.exceptions.PreventUpdate

    try:
        # Use the helper to generate export
        download_data = export_helper.export_devices(format=export_format)

        if download_data:
            toast = ToastManager.success(
                "Export Complete",
                detail_message=f"Devices exported as {export_format.upper()}"
            )
            return download_data, toast
        else:
            toast = ToastManager.error(
                "Export Failed",
                detail_message="Could not generate export file"
            )
            return None, toast

    except Exception as e:
        logger.error(f"Error exporting devices: {e}")
        toast = ToastManager.error(
            "Export Error",
            detail_message=str(e)
        )
        return None, toast


4. Add export format dropdown to your UI:

dbc.Select(
    id='export-format-dropdown',
    options=[
        {'label': 'CSV Format', 'value': 'csv'},
        {'label': 'JSON Format', 'value': 'json'},
        {'label': 'PDF Report', 'value': 'pdf'},
        {'label': 'Excel Workbook', 'value': 'excel'}
    ],
    value='csv',
    className='mb-2'
)

dbc.Button(
    "Export Devices",
    id='export-devices-btn',
    color="primary"
)

dcc.Download(id='download-devices')


5. For alerts export with time range:

@app.callback(
    [Output('download-alerts', 'data'),
     Output('alerts-export-toast', 'children')],
    [Input('export-alerts-btn', 'n_clicks')],
    [State('export-format-dropdown', 'value'),
     State('alert-days-input', 'value')]
)
def export_alerts_universal(n_clicks, export_format, days):
    '''Export alerts in selected format with time range.'''
    if not n_clicks:
        raise dash.exceptions.PreventUpdate

    try:
        days = int(days) if days else 7

        download_data = export_helper.export_alerts(
            format=export_format,
            days=days
        )

        if download_data:
            toast = ToastManager.success(
                "Export Complete",
                detail_message=f"Last {days} days of alerts exported as {export_format.upper()}"
            )
            return download_data, toast
        else:
            toast = ToastManager.error(
                "Export Failed",
                detail_message="Could not generate export file"
            )
            return None, toast

    except Exception as e:
        logger.error(f"Error exporting alerts: {e}")
        toast = ToastManager.error(
            "Export Error",
            detail_message=str(e)
        )
        return None, toast


6. For connections export with device filter:

@app.callback(
    [Output('download-connections', 'data'),
     Output('connections-export-toast', 'children')],
    [Input('export-connections-btn', 'n_clicks')],
    [State('export-format-dropdown', 'value'),
     State('device-selector', 'value'),
     State('hours-input', 'value')]
)
def export_connections_universal(n_clicks, export_format, device_ip, hours):
    '''Export connections in selected format with filters.'''
    if not n_clicks:
        raise dash.exceptions.PreventUpdate

    try:
        hours = int(hours) if hours else 24

        download_data = export_helper.export_connections(
            format=export_format,
            device_ip=device_ip,
            hours=hours
        )

        if download_data:
            toast = ToastManager.success(
                "Export Complete",
                detail_message=f"Connections exported as {export_format.upper()}"
            )
            return download_data, toast
        else:
            toast = ToastManager.error(
                "Export Failed",
                detail_message="Could not generate export file"
            )
            return None, toast

    except Exception as e:
        logger.error(f"Error exporting connections: {e}")
        toast = ToastManager.error(
            "Export Error",
            detail_message=str(e)
        )
        return None, toast
"""
