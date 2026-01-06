#!/usr/bin/env python3
"""
Report Generator for IoTSentinel

Provides export functionality for devices, alerts, and connections.
Generates PDF reports with charts and statistics.
Enhanced with trend analysis and advanced visualizations.
"""

import csv
import sqlite3
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from pathlib import Path
import io
import json
from io import BytesIO

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates reports and exports data from IoTSentinel database.

    Enhanced with trend analysis, executive summaries, and advanced visualizations.
    """

    def __init__(self, db_path: str):
        """
        Initialize report generator.

        Args:
            db_path: Path to SQLite database
        """
        self.db_path = db_path

        # Lazy load heavy dependencies
        self._trend_analyzer = None
        self._pdf_exporter = None
        self._excel_exporter = None

    @property
    def trend_analyzer(self):
        """Lazy load TrendAnalyzer."""
        if self._trend_analyzer is None:
            try:
                from .trend_analyzer import TrendAnalyzer
                self._trend_analyzer = TrendAnalyzer(self.db_path)
            except ImportError:
                logger.warning("TrendAnalyzer not available")
        return self._trend_analyzer

    @property
    def pdf_exporter(self):
        """Lazy load PDFExporter."""
        if self._pdf_exporter is None:
            try:
                from .pdf_exporter import PDFExporter
                self._pdf_exporter = PDFExporter(self.db_path)
            except ImportError:
                logger.warning("PDFExporter not available")
        return self._pdf_exporter

    @property
    def excel_exporter(self):
        """Lazy load ExcelExporter."""
        if self._excel_exporter is None:
            try:
                from .excel_exporter import ExcelExporter
                self._excel_exporter = ExcelExporter(self.db_path)
            except ImportError:
                logger.warning("ExcelExporter not available")
        return self._excel_exporter

    def export_devices_csv(self) -> str:
        """
        Export all devices to CSV format.

        Returns:
            CSV string with device data
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    device_ip,
                    device_name,
                    device_type,
                    mac_address,
                    manufacturer,
                    first_seen,
                    last_seen,
                    is_trusted,
                    is_blocked
                FROM devices
                ORDER BY last_seen DESC
            """)

            devices = cursor.fetchall()
            conn.close()

            # Create CSV in memory
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                'IP Address', 'Device Name', 'Type', 'MAC Address',
                'Manufacturer', 'First Seen', 'Last Seen', 'Trusted', 'Blocked'
            ])

            # Write data rows
            for device in devices:
                writer.writerow([
                    device['device_ip'],
                    device['device_name'] or 'Unknown',
                    device['device_type'] or 'Unknown',
                    device['mac_address'] or 'Unknown',
                    device['manufacturer'] or 'Unknown',
                    device['first_seen'],
                    device['last_seen'],
                    'Yes' if device['is_trusted'] else 'No',
                    'Yes' if device['is_blocked'] else 'No'
                ])

            csv_data = output.getvalue()
            output.close()

            logger.info(f"Exported {len(devices)} devices to CSV")
            return csv_data

        except sqlite3.Error as e:
            logger.error(f"Database error exporting devices: {e}")
            return ""

    def export_alerts_csv(self, days: int = 7) -> str:
        """
        Export alerts to CSV format.

        Args:
            days: Number of days to export (default: 7)

        Returns:
            CSV string with alert data
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cutoff_date = datetime.now() - timedelta(days=days)

            cursor.execute("""
                SELECT
                    a.id,
                    a.timestamp,
                    a.device_ip,
                    d.device_name,
                    a.severity,
                    a.anomaly_score,
                    a.explanation,
                    a.acknowledged
                FROM alerts a
                LEFT JOIN devices d ON a.device_ip = d.device_ip
                WHERE a.timestamp > ?
                ORDER BY a.timestamp DESC
            """, (cutoff_date.isoformat(),))

            alerts = cursor.fetchall()
            conn.close()

            # Create CSV in memory
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                'Alert ID', 'Timestamp', 'Device IP', 'Device Name',
                'Severity', 'Anomaly Score', 'Explanation', 'Acknowledged'
            ])

            # Write data rows
            for alert in alerts:
                writer.writerow([
                    alert['id'],
                    alert['timestamp'],
                    alert['device_ip'],
                    alert['device_name'] or 'Unknown',
                    alert['severity'].upper(),
                    f"{alert['anomaly_score']:.4f}" if alert['anomaly_score'] else 'N/A',
                    alert['explanation'] or '',
                    'Yes' if alert['acknowledged'] else 'No'
                ])

            csv_data = output.getvalue()
            output.close()

            logger.info(f"Exported {len(alerts)} alerts from last {days} days to CSV")
            return csv_data

        except sqlite3.Error as e:
            logger.error(f"Database error exporting alerts: {e}")
            return ""

    def export_connections_csv(self, device_ip: Optional[str] = None, hours: int = 24) -> str:
        """
        Export connection logs to CSV format.

        Args:
            device_ip: Filter by specific device IP (optional)
            hours: Number of hours to export (default: 24)

        Returns:
            CSV string with connection data
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cutoff_time = datetime.now() - timedelta(hours=hours)

            if device_ip:
                cursor.execute("""
                    SELECT
                        timestamp,
                        device_ip,
                        dest_ip,
                        dest_port,
                        protocol,
                        service,
                        duration,
                        bytes_sent,
                        bytes_received,
                        packets_sent,
                        packets_received,
                        conn_state
                    FROM connections
                    WHERE device_ip = ? AND timestamp > ?
                    ORDER BY timestamp DESC
                """, (device_ip, cutoff_time.isoformat()))
            else:
                cursor.execute("""
                    SELECT
                        timestamp,
                        device_ip,
                        dest_ip,
                        dest_port,
                        protocol,
                        service,
                        duration,
                        bytes_sent,
                        bytes_received,
                        packets_sent,
                        packets_received,
                        conn_state
                    FROM connections
                    WHERE timestamp > ?
                    ORDER BY timestamp DESC
                    LIMIT 10000
                """, (cutoff_time.isoformat(),))

            connections = cursor.fetchall()
            conn.close()

            # Create CSV in memory
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                'Timestamp', 'Source IP', 'Destination IP', 'Destination Port',
                'Protocol', 'Service', 'Duration (s)', 'Bytes Sent', 'Bytes Received',
                'Packets Sent', 'Packets Received', 'Connection State'
            ])

            # Write data rows
            for conn in connections:
                writer.writerow([
                    conn['timestamp'],
                    conn['device_ip'],
                    conn['dest_ip'] or 'N/A',
                    conn['dest_port'] or 'N/A',
                    conn['protocol'] or 'N/A',
                    conn['service'] or 'Unknown',
                    f"{conn['duration']:.2f}" if conn['duration'] else 'N/A',
                    conn['bytes_sent'] or 0,
                    conn['bytes_received'] or 0,
                    conn['packets_sent'] or 0,
                    conn['packets_received'] or 0,
                    conn['conn_state'] or 'Unknown'
                ])

            csv_data = output.getvalue()
            output.close()

            logger.info(f"Exported {len(connections)} connections to CSV")
            return csv_data

        except sqlite3.Error as e:
            logger.error(f"Database error exporting connections: {e}")
            return ""

    def get_summary_statistics(self, days: int = 7) -> Dict[str, Any]:
        """
        Get summary statistics for report generation.

        Args:
            days: Number of days to analyze

        Returns:
            Dictionary with statistics
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cutoff_date = datetime.now() - timedelta(days=days)

            stats = {}

            # Total devices
            cursor.execute("SELECT COUNT(*) FROM devices")
            stats['total_devices'] = cursor.fetchone()[0]

            # Active devices (seen in last 24 hours)
            active_cutoff = datetime.now() - timedelta(hours=24)
            cursor.execute("SELECT COUNT(*) FROM devices WHERE last_seen > ?", (active_cutoff.isoformat(),))
            stats['active_devices'] = cursor.fetchone()[0]

            # Blocked devices
            cursor.execute("SELECT COUNT(*) FROM devices WHERE is_blocked = 1")
            stats['blocked_devices'] = cursor.fetchone()[0]

            # Total alerts in period
            cursor.execute("SELECT COUNT(*) FROM alerts WHERE timestamp > ?", (cutoff_date.isoformat(),))
            stats['total_alerts'] = cursor.fetchone()[0]

            # Alerts by severity
            cursor.execute("""
                SELECT severity, COUNT(*) as count
                FROM alerts
                WHERE timestamp > ?
                GROUP BY severity
            """, (cutoff_date.isoformat(),))
            stats['alerts_by_severity'] = dict(cursor.fetchall())

            # Total connections in period
            cursor.execute("SELECT COUNT(*) FROM connections WHERE timestamp > ?", (cutoff_date.isoformat(),))
            stats['total_connections'] = cursor.fetchone()[0]

            # Total data transferred (MB)
            cursor.execute("""
                SELECT SUM(bytes_sent + bytes_received) as total_bytes
                FROM connections
                WHERE timestamp > ?
            """, (cutoff_date.isoformat(),))
            result = cursor.fetchone()
            stats['total_data_mb'] = (result[0] / (1024 * 1024)) if result[0] else 0

            # Top talkers (devices by data volume)
            cursor.execute("""
                SELECT
                    c.device_ip,
                    d.device_name,
                    SUM(c.bytes_sent + c.bytes_received) as total_bytes
                FROM connections c
                LEFT JOIN devices d ON c.device_ip = d.device_ip
                WHERE c.timestamp > ?
                GROUP BY c.device_ip
                ORDER BY total_bytes DESC
                LIMIT 10
            """, (cutoff_date.isoformat(),))
            stats['top_talkers'] = [
                {
                    'ip': row[0],
                    'name': row[1] or 'Unknown',
                    'data_mb': (row[2] / (1024 * 1024)) if row[2] else 0
                }
                for row in cursor.fetchall()
            ]

            # Most common protocols
            cursor.execute("""
                SELECT protocol, COUNT(*) as count
                FROM connections
                WHERE timestamp > ? AND protocol IS NOT NULL
                GROUP BY protocol
                ORDER BY count DESC
                LIMIT 10
            """, (cutoff_date.isoformat(),))
            stats['top_protocols'] = dict(cursor.fetchall())

            # Most contacted destinations
            cursor.execute("""
                SELECT dest_ip, COUNT(*) as count
                FROM connections
                WHERE timestamp > ? AND dest_ip IS NOT NULL
                GROUP BY dest_ip
                ORDER BY count DESC
                LIMIT 10
            """, (cutoff_date.isoformat(),))
            stats['top_destinations'] = dict(cursor.fetchall())

            conn.close()

            logger.info(f"Generated summary statistics for {days} days")
            return stats

        except sqlite3.Error as e:
            logger.error(f"Database error generating statistics: {e}")
            return {}

    def export_alert_rules_csv(self) -> str:
        """
        Export alert rules to CSV format.

        Returns:
            CSV string with alert rule data
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    id,
                    name,
                    description,
                    rule_type,
                    condition_operator,
                    threshold_value,
                    time_window_hours,
                    severity,
                    is_enabled,
                    trigger_count,
                    last_triggered,
                    created_at
                FROM alert_rules
                ORDER BY severity DESC, name ASC
            """)

            rules = cursor.fetchall()
            conn.close()

            # Create CSV in memory
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                'Rule ID', 'Name', 'Description', 'Type', 'Condition',
                'Threshold', 'Time Window (hrs)', 'Severity', 'Enabled',
                'Trigger Count', 'Last Triggered', 'Created At'
            ])

            # Write data rows
            for rule in rules:
                writer.writerow([
                    rule['id'],
                    rule['name'],
                    rule['description'] or '',
                    rule['rule_type'],
                    rule['condition_operator'],
                    rule['threshold_value'] or 'N/A',
                    rule['time_window_hours'],
                    rule['severity'].upper(),
                    'Yes' if rule['is_enabled'] else 'No',
                    rule['trigger_count'] or 0,
                    rule['last_triggered'] or 'Never',
                    rule['created_at']
                ])

            csv_data = output.getvalue()
            output.close()

            logger.info(f"Exported {len(rules)} alert rules to CSV")
            return csv_data

        except sqlite3.Error as e:
            logger.error(f"Database error exporting alert rules: {e}")
            return ""

    def generate_executive_summary(self, days: int = 7) -> str:
        """
        Generate executive summary text for reports.

        Args:
            days: Number of days to analyze

        Returns:
            Formatted summary text
        """
        stats = self.get_summary_statistics(days)

        if not stats:
            return "Unable to generate summary - no data available."

        summary_lines = [
            f"# IoTSentinel Security Report",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Report Period: Last {days} days",
            "",
            "## Network Overview",
            f"- Total Devices: {stats.get('total_devices', 0)}",
            f"- Active Devices (24h): {stats.get('active_devices', 0)}",
            f"- Blocked Devices: {stats.get('blocked_devices', 0)}",
            "",
            "## Security Alerts",
            f"- Total Alerts: {stats.get('total_alerts', 0)}",
        ]

        # Add severity breakdown
        alerts_by_sev = stats.get('alerts_by_severity', {})
        if alerts_by_sev:
            summary_lines.append("- Alerts by Severity:")
            for severity in ['critical', 'high', 'medium', 'low']:
                count = alerts_by_sev.get(severity, 0)
                if count > 0:
                    summary_lines.append(f"  - {severity.upper()}: {count}")

        summary_lines.extend([
            "",
            "## Network Activity",
            f"- Total Connections: {stats.get('total_connections', 0):,}",
            f"- Data Transferred: {stats.get('total_data_mb', 0):.2f} MB",
        ])

        # Add top protocols
        top_protocols = stats.get('top_protocols', {})
        if top_protocols:
            summary_lines.append("- Top Protocols:")
            for protocol, count in list(top_protocols.items())[:5]:
                summary_lines.append(f"  - {protocol}: {count:,} connections")

        # Add top talkers
        top_talkers = stats.get('top_talkers', [])
        if top_talkers:
            summary_lines.append("")
            summary_lines.append("## Top Data Consumers")
            for i, talker in enumerate(top_talkers[:5], 1):
                summary_lines.append(f"{i}. {talker['name']} ({talker['ip']}): {talker['data_mb']:.2f} MB")

        return "\n".join(summary_lines)

    # ==================== Enhanced Reporting Methods ====================

    def generate_executive_summary_report(
        self,
        days: int = 7,
        format: str = 'pdf'
    ) -> Optional[Dict[str, Any]]:
        """
        Generate comprehensive executive summary report with trends.

        Args:
            days: Number of days to analyze
            format: Output format ('pdf', 'json', 'excel')

        Returns:
            Dictionary with report content and metadata
        """
        try:
            if not self.trend_analyzer:
                logger.error("TrendAnalyzer not available")
                return None

            # Get executive summary data with trends
            summary = self.trend_analyzer.get_executive_summary(days=days)

            if format == 'json':
                return {
                    'content': json.dumps(summary, indent=2),
                    'filename': f'executive_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',
                    'mimetype': 'application/json'
                }

            elif format == 'pdf':
                return self._generate_executive_summary_pdf(summary, days)

            elif format == 'excel':
                return self._generate_executive_summary_excel(summary, days)

            else:
                logger.error(f"Unsupported format: {format}")
                return None

        except Exception as e:
            logger.error(f"Error generating executive summary: {e}")
            return None

    def generate_trend_analysis_report(
        self,
        metric: str = 'alerts',
        days: int = 30,
        granularity: str = 'daily',
        format: str = 'json'
    ) -> Optional[Dict[str, Any]]:
        """
        Generate detailed trend analysis report.

        Args:
            metric: Metric to analyze ('alerts', 'connections', 'devices')
            days: Number of days to analyze
            granularity: Time granularity ('hourly', 'daily', 'weekly')
            format: Output format ('pdf', 'json')

        Returns:
            Dictionary with report content and metadata
        """
        try:
            if not self.trend_analyzer:
                logger.error("TrendAnalyzer not available")
                return None

            if metric == 'alerts':
                trend_data = self.trend_analyzer.analyze_alert_trends(
                    days=days,
                    granularity=granularity
                )
            elif metric == 'connections':
                trend_data = self.trend_analyzer.analyze_network_traffic(
                    hours=days * 24
                )
            elif metric == 'devices':
                trend_data = self.trend_analyzer.analyze_device_activity(days=days)
            else:
                logger.error(f"Unknown metric: {metric}")
                return None

            return {
                'content': json.dumps(trend_data, indent=2),
                'filename': f'{metric}_trend_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',
                'mimetype': 'application/json'
            }

        except Exception as e:
            logger.error(f"Error generating trend analysis report: {e}")
            return None

    def generate_security_posture_report(
        self,
        days: int = 30,
        format: str = 'json'
    ) -> Optional[Dict[str, Any]]:
        """
        Generate comprehensive security posture report.

        Args:
            days: Number of days to analyze
            format: Output format ('pdf', 'json', 'excel')

        Returns:
            Dictionary with report content and metadata
        """
        try:
            if not self.trend_analyzer:
                logger.error("TrendAnalyzer not available")
                return None

            # Gather all security metrics
            alert_trends = self.trend_analyzer.analyze_alert_trends(days=days)
            device_activity = self.trend_analyzer.analyze_device_activity(days=days)
            network_traffic = self.trend_analyzer.analyze_network_traffic(hours=days*24)
            anomalies = self.trend_analyzer.detect_anomalies(metric='alerts', days=days)

            security_data = {
                'alert_trends': alert_trends,
                'device_activity': device_activity,
                'network_traffic': network_traffic,
                'anomalies': anomalies,
                'period_days': days,
                'generated_at': datetime.now().isoformat()
            }

            return {
                'content': json.dumps(security_data, indent=2),
                'filename': f'security_posture_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',
                'mimetype': 'application/json'
            }

        except Exception as e:
            logger.error(f"Error generating security posture report: {e}")
            return None

    def generate_chart_data(
        self,
        chart_type: str,
        data_source: str,
        days: int = 7
    ) -> Optional[Dict[str, Any]]:
        """
        Generate chart data for dashboard integration.

        Args:
            chart_type: Type of chart ('trend', 'bar', 'pie', 'area')
            data_source: Data source ('alerts', 'connections', 'devices')
            days: Number of days to analyze

        Returns:
            Plotly figure dictionary ready for dcc.Graph
        """
        try:
            from .chart_factory import ChartFactory

            if not self.trend_analyzer:
                return ChartFactory.create_empty_chart('Trend analyzer not available')

            if data_source == 'alerts':
                trend_data = self.trend_analyzer.analyze_alert_trends(
                    days=days,
                    granularity='daily'
                )

                if chart_type == 'trend':
                    time_series = trend_data['time_series']
                    if time_series:
                        x_vals = [t[0] for t in time_series]
                        y_vals = [t[1] for t in time_series]
                        return ChartFactory.create_trend_chart(
                            x_vals, y_vals,
                            title=f'Alert Trends - Last {days} Days',
                            x_title='Date',
                            y_title='Alert Count'
                        )

                elif chart_type == 'area':
                    time_series = trend_data['time_series']
                    if time_series:
                        x_vals = [t[0] for t in time_series]
                        y_vals = [t[1] for t in time_series]
                        return ChartFactory.create_area_chart(
                            x_vals, y_vals,
                            title=f'Alert Volume - Last {days} Days',
                            x_title='Date',
                            y_title='Alerts'
                        )

            elif data_source == 'devices':
                activity_data = self.trend_analyzer.analyze_device_activity(days=days)

                if chart_type == 'bar':
                    devices = activity_data['most_active_devices'][:10]
                    if devices:
                        ips = [d[0] for d in devices]
                        counts = [d[1] for d in devices]
                        return ChartFactory.create_bar_chart(
                            ips, counts,
                            title='Top 10 Most Active Devices',
                            x_title='Device IP',
                            y_title='Connection Count',
                            tick_angle=-30
                        )

            return ChartFactory.create_empty_chart('No data available')

        except Exception as e:
            logger.error(f"Error generating chart data: {e}")
            try:
                from .chart_factory import ChartFactory
                return ChartFactory.create_empty_chart(f'Error: {str(e)}')
            except:
                return None

    def _generate_executive_summary_pdf(
        self,
        summary: Dict[str, Any],
        days: int
    ) -> Optional[Dict[str, Any]]:
        """
        Generate PDF executive summary report with charts.

        Args:
            summary: Summary data from trend analyzer
            days: Analysis period in days

        Returns:
            PDF report as dictionary
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#0d6efd'),
                spaceAfter=30
            )
            story.append(Paragraph('Executive Summary Report', title_style))
            story.append(Spacer(1, 0.2*inch))

            # Period information
            period = summary.get('period', {})
            story.append(Paragraph(
                f"<b>Report Period:</b> {period.get('start_date', 'N/A')} to {period.get('end_date', 'N/A')}",
                styles['Normal']
            ))
            story.append(Spacer(1, 0.3*inch))

            # Security Posture Section
            story.append(Paragraph('<b>Security Posture</b>', styles['Heading2']))
            security = summary.get('security_posture', {})
            security_data = [
                ['Metric', 'Value'],
                ['Alert Trend', security.get('alert_trend', 'N/A').upper()],
                ['Total Alerts', str(security.get('total_alerts', 0))],
                ['Critical Alerts', str(security.get('critical_alerts', 0))],
                ['Trend Change', f"{security.get('percent_change', 0):+.1f}%"]
            ]
            security_table = Table(security_data, colWidths=[3*inch, 2*inch])
            security_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(security_table)
            story.append(Spacer(1, 0.3*inch))

            # Network Activity Section
            story.append(Paragraph('<b>Network Activity</b>', styles['Heading2']))
            network = summary.get('network_activity', {})
            network_data = [
                ['Metric', 'Value'],
                ['Total Connections', f"{network.get('total_connections', 0):,}"],
                ['Blocked Connections', str(network.get('blocked_connections', 0))],
                ['Suspicious Patterns', str(network.get('suspicious_patterns', 0))],
                ['Unique Sources', str(network.get('unique_sources', 0))]
            ]
            network_table = Table(network_data, colWidths=[3*inch, 2*inch])
            network_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#17a2b8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(network_table)
            story.append(Spacer(1, 0.3*inch))

            # Device Status Section
            story.append(Paragraph('<b>Device Status</b>', styles['Heading2']))
            devices = summary.get('device_status', {})
            device_data = [
                ['Metric', 'Value'],
                ['Active Devices', str(devices.get('active_devices', 0))],
                ['New Devices', str(devices.get('new_devices', 0))],
                ['Inactive Devices', str(devices.get('inactive_devices', 0))]
            ]
            device_table = Table(device_data, colWidths=[3*inch, 2*inch])
            device_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(device_table)
            story.append(Spacer(1, 0.3*inch))

            # Top Concerns Section
            story.append(Paragraph('<b>Top Concerns</b>', styles['Heading2']))
            concerns = summary.get('top_concerns', [])
            for i, concern in enumerate(concerns, 1):
                story.append(Paragraph(f"{i}. {concern}", styles['Normal']))
                story.append(Spacer(1, 0.1*inch))

            # Build PDF
            doc.build(story)
            pdf_content = buffer.getvalue()
            buffer.close()

            return {
                'content': pdf_content,
                'filename': f'executive_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
                'mimetype': 'application/pdf'
            }

        except Exception as e:
            logger.error(f"Error generating executive summary PDF: {e}")
            return None

    def _generate_executive_summary_excel(
        self,
        summary: Dict[str, Any],
        days: int
    ) -> Optional[Dict[str, Any]]:
        """
        Generate Excel executive summary report.

        Args:
            summary: Summary data
            days: Analysis period

        Returns:
            Excel report as dictionary
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Executive Summary"

            # Title
            ws['A1'] = 'Executive Summary Report'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:B1')

            # Period
            period = summary.get('period', {})
            ws['A3'] = 'Report Period:'
            ws['B3'] = f"{period.get('start_date', 'N/A')} to {period.get('end_date', 'N/A')}"

            # Security Posture
            row = 5
            ws[f'A{row}'] = 'Security Posture'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            security = summary.get('security_posture', {})
            metrics = [
                ('Alert Trend', security.get('alert_trend', 'N/A').upper()),
                ('Total Alerts', security.get('total_alerts', 0)),
                ('Critical Alerts', security.get('critical_alerts', 0)),
                ('Trend Change', f"{security.get('percent_change', 0):+.1f}%")
            ]

            for metric, value in metrics:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                row += 1

            # Network Activity
            row += 1
            ws[f'A{row}'] = 'Network Activity'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            network = summary.get('network_activity', {})
            metrics = [
                ('Total Connections', f"{network.get('total_connections', 0):,}"),
                ('Blocked Connections', network.get('blocked_connections', 0)),
                ('Suspicious Patterns', network.get('suspicious_patterns', 0)),
                ('Unique Sources', network.get('unique_sources', 0))
            ]

            for metric, value in metrics:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                row += 1

            # Auto-size columns
            for col in ['A', 'B']:
                ws.column_dimensions[col].width = 30

            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()

            return {
                'content': excel_content,
                'filename': f'executive_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }

        except Exception as e:
            logger.error(f"Error generating executive summary Excel: {e}")
            return None


def save_report_to_file(content: str, filename: str, report_dir: str = "data/reports") -> str:
    """
    Save report content to file.

    Args:
        content: Report content
        filename: Filename to save
        report_dir: Directory to save reports

    Returns:
        Full path to saved file
    """
    try:
        # Create reports directory
        report_path = Path(report_dir)
        report_path.mkdir(parents=True, exist_ok=True)

        # Generate timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        full_filename = f"{timestamp}_{filename}"
        file_path = report_path / full_filename

        # Write content
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)

        logger.info(f"Saved report to {file_path}")
        return str(file_path)

    except Exception as e:
        logger.error(f"Error saving report: {e}")
        return ""
